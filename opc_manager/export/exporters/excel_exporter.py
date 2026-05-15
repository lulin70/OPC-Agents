import io

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    def export(self, data, template=None, **opts) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = opts.get("sheet_name", "Report")

        title = data.metadata.get("title", "Report")
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        rows = self._parse_markdown_table(data.content)
        start_row = 3
        for r_idx, row in enumerate(rows):
            for c_idx, cell in enumerate(row):
                ws.cell(row=start_row + r_idx, column=c_idx + 1, value=cell)
                if r_idx == 0:
                    ws.cell(row=start_row + r_idx, column=c_idx + 1).font = Font(bold=True)
                    ws.cell(row=start_row + r_idx, column=c_idx + 1).fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")

        for col in range(1, max(len(r) for r in rows) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _parse_markdown_table(self, content):
        rows = []
        in_table = False
        for line in content.split('\n'):
            stripped = line.strip()
            if stripped.startswith('|'):
                cells = [c.strip() for c in stripped.split('|')]
                cells = [c for c in cells if c]
                if all(c.replace('-', '').replace(':', '') == '' for c in cells):
                    continue
                in_table = True
                rows.append(cells)
            elif in_table and stripped == '':
                break
            elif not in_table:
                if stripped and not stripped.startswith('#'):
                    rows.append([stripped])
        if not rows:
            rows = [[content]]
        return rows
